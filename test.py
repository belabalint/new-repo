from tkinter import filedialog
from tkinter import Tk
from pandas import read_table, DataFrame
from numpy import arange, pi
import dft
from openpyxl.formatting import Rule
from openpyxl.formatting.rule import FormulaRule, CellIsRule
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.styles import Color, PatternFill
twopi = pi * 2


#getting filepath with gui
root = Tk()
root.withdraw()
filepath = filedialog.askopenfilename()

df = read_table(filepath, delimiter=' ', skiprows=1)
times = df['time[s]'].to_numpy()
df = df.drop("Cz", axis=1)
df = df.iloc[:, 1:19]
fp1 = df['Fp1'].to_numpy()
cols = df.columns.tolist()
cols = ['Fp1', 'Fp2', 'F7', 'F8', 'C3', 'C4', 'F3', 'F4', 'Fz', 'T3', 'T4', 'P3', 'Pz', 'P4', 'O1', 'O2', 'T5', 'T6']
df = df[cols]



#verify deltat
deltat = round(times[1] - times[0], 10)
if not (deltat == 0.002 or deltat == 0.004):
    print('incorrect deltaT')
    raise ValueError
for t in arange(1, len(times)):
    if round(times[t] - times[t-1], 6) != deltat:
        print('incorrect file, inconsistent deltaT')
        raise ValueError

#find 0 timepoint
index = dft.continuoustransform(fp1, times, deltat)
print(index)

#check data
l = [index]+[index + 30000 + i * 1500 for i in range(37)]
dfs = [df.iloc[l[i]:l[i+1]] for i in range(len(l)-1)]


data = [[dft.checkdata(testindex, electrodeindex, dfs[testindex].iloc[:, electrodeindex]) for electrodeindex in range(18)] for testindex in range(37)]
cols = ['fp1', 'fp2', 'f7', 'f3', 'fz', 'f4', 'f8', 't3', 'c3', 'c4', 't4', 't5', 'p3', 'pz', 'p4', 't6', 'o1', 'o2']
cols1 = ['fp1', 'fp2', 'f7', 'f8', 'c3', 'c4', 'f3', 'f4', 'fz', 't3', 't4', 'p3', 'pz', 'p4', 'o1', 'o2', 't5', 't6']
rows = ['noise test']
for label in cols1:
    rows += [f'{label} 5hz no resistance', f'{label} 5hz with resistance']
scores_df = DataFrame(data, columns=cols1, index=rows)

#strlist = filepath.split("/")
strlist = filepath.split(".")
file_name = f'{strlist[0]}.xlsx'
scores_df.to_excel(file_name, 'sheet1', )


wb = load_workbook(filename = file_name)
sheet = wb['sheet1']

red_fill = PatternFill(bgColor="FF0000")
green_fill = PatternFill(bgColor="00FF00")


dxf = DifferentialStyle(fill=red_fill)
rule = Rule(type="containsText", operator="containsText", text="highlight", dxf=dxf)
rule.formula = ['NOT(ISERROR(SEARCH("FAILED",A1)))']
sheet.conditional_formatting.add('A1:S38', rule)
dxf = DifferentialStyle(fill=green_fill)
rule = Rule(type="containsText", operator="containsText", text="highlight", dxf=dxf)
rule.formula = ['NOT(ISERROR(SEARCH("passed",A1)))']
sheet.conditional_formatting.add('A1:S38', rule)
wb.save(file_name)